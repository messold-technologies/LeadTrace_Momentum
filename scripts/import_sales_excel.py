"""
Sales Excel -> MongoDB importer.
Auto-detects phone / date / center columns — no hardcoded sheet config needed.
Works for any number of channels / sheets.
"""

import re
import sys
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from pymongo import MongoClient, UpdateOne
from pymongo.errors import BulkWriteError

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ── tunables ──────────────────────────────────────────────────────────────────
MONGO_URI  = "mongodb://localhost:27017"
DB_NAME    = "leadtrace"
COLLECTION = "sales"

# Sheets that are NOT sales channels — always skip
NON_SALES_SHEETS = {"DNC List", "Sheet1"}

# Minimum detection score to accept a column as phone / date / center
MIN_PHONE_SCORE  = 5
MIN_DATE_SCORE   = 5
MIN_CENTER_SCORE = 8   # stricter — avoids false positives like "SALE DATE"

# Header keyword lists (substring match, lower-cased)
PHONE_KEYWORDS  = ["mobile", "phone", "contact", " no", "number", "ph ", "mob"]
DATE_KEYWORDS   = ["date", "agreement", "doa", "sold"]
CENTER_KEYWORDS = ["center", "centre", "branch", "hub", "location"]
# ─────────────────────────────────────────────────────────────────────────────

PHONE_RE = re.compile(r"^\d{9,10}$")


def normalize_phone(val) -> "str | None":
    if val is None:
        return None
    s = re.sub(r"[\s\-\(\)\+]", "", str(val))
    if PHONE_RE.match(s):
        # Australian mobiles are often stored without leading 0 (9 digits)
        return s if len(s) == 10 else "0" + s
    return None


def parse_date(val) -> "datetime | None":
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                pass
    return None


def looks_like_phone(val) -> bool:
    return normalize_phone(val) is not None


def looks_like_date(val) -> bool:
    return parse_date(val) is not None


def find_header_row(ws, max_scan: int = 10) -> int:
    """Return 1-based row index with the most string values — that's the header row."""
    best_row, best_score = 1, 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        score = sum(1 for v in row if isinstance(v, str) and len(v.strip()) > 1)
        if score > best_score:
            best_score, best_row = score, i
    return best_row


def detect_columns(headers: list, sample_rows: list) -> dict:
    """
    Score every column for phone / date / center role using header keywords
    AND actual value patterns. Returns best column index for each role.
    """
    phone_best  = (None, 0)   # (col_idx, score)
    date_best   = (None, 0)
    center_best = (None, 0)

    for idx, h in enumerate(headers):
        h_lower = (h or "").lower().strip()
        vals = [r[idx] for r in sample_rows if idx < len(r) and r[idx] is not None]

        # ── phone: keyword weight × 3 + value pattern matches ──
        ph = sum(1 for kw in PHONE_KEYWORDS if kw in h_lower) * 3
        ph += sum(1 for v in vals[:15] if looks_like_phone(v))
        if ph > phone_best[1]:
            phone_best = (idx, ph)

        # ── date: keyword weight × 3 + value pattern matches ──
        dt = sum(1 for kw in DATE_KEYWORDS if kw in h_lower) * 3
        dt += sum(1 for v in vals[:15] if looks_like_date(v))
        if dt > date_best[1]:
            date_best = (idx, dt)

        # ── center: keyword weight × 4 + string value heuristic ──
        ct = sum(1 for kw in CENTER_KEYWORDS if kw in h_lower) * 4
        ct += sum(
            1 for v in vals[:15]
            if isinstance(v, str) and 2 < len(v) < 80
            and not looks_like_date(v) and not looks_like_phone(v)
        )
        if ct > center_best[1]:
            center_best = (idx, ct)

    return {
        "phone":  phone_best[0]  if phone_best[1]  >= MIN_PHONE_SCORE  else None,
        "date":   date_best[0]   if date_best[1]   >= MIN_DATE_SCORE   else None,
        "center": center_best[0] if center_best[1] >= MIN_CENTER_SCORE else None,
    }


def import_excel(filepath: str, mongo_uri: str = MONGO_URI) -> list:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    client = MongoClient(mongo_uri)
    col = client[DB_NAME][COLLECTION]

    col.create_index("phone")
    col.create_index([("phone", 1), ("channel", 1)])

    report = []

    for sheet_name in wb.sheetnames:
        if sheet_name in NON_SALES_SHEETS:
            log.info("Skipping non-sales sheet: %s", sheet_name)
            continue

        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        all_rows   = list(ws.iter_rows(min_row=1, values_only=True))

        if header_row > len(all_rows):
            log.warning("Sheet %r: no data rows", sheet_name)
            continue

        headers   = [str(v).strip() if v else "" for v in all_rows[header_row - 1]]
        data_rows = all_rows[header_row:]
        cols      = detect_columns(headers, data_rows[:20])

        if cols["phone"] is None:
            log.warning("Sheet %r: no phone column detected — skipping", sheet_name)
            report.append({"sheet": sheet_name, "skipped": True, "reason": "no phone column"})
            continue

        log.info(
            "Sheet %r | header_row=%d | phone=%r | date=%r | center=%r",
            sheet_name, header_row,
            headers[cols["phone"]],
            headers[cols["date"]]   if cols["date"]   is not None else None,
            headers[cols["center"]] if cols["center"] is not None else None,
        )

        ops, skipped_rows = [], 0

        for row in data_rows:
            raw_phone = row[cols["phone"]] if cols["phone"] < len(row) else None
            phone = normalize_phone(raw_phone)
            if not phone:
                skipped_rows += 1
                continue

            sale_date = parse_date(
                row[cols["date"]] if cols["date"] is not None and cols["date"] < len(row) else None
            )
            center_name = (
                row[cols["center"]] if cols["center"] is not None and cols["center"] < len(row) else None
            )
            if isinstance(center_name, str):
                center_name = center_name.strip() or None

            doc = {
                "phone":       phone,
                "channel":     sheet_name,
                "sale_date":   sale_date,
                "center_name": center_name,
                "imported_at": datetime.utcnow(),
            }
            # same phone + channel + date = same sale record (prevents duplicates on re-import)
            ops.append(UpdateOne(
                {"phone": phone, "channel": sheet_name, "sale_date": sale_date},
                {"$setOnInsert": doc},
                upsert=True,
            ))

        if ops:
            try:
                result = col.bulk_write(ops, ordered=False)
                inserted, matched = result.upserted_count, result.matched_count
            except BulkWriteError as e:
                inserted = e.details.get("nUpserted", 0)
                matched  = e.details.get("nMatched", 0)
                log.warning("Sheet %r partial error: %s", sheet_name, e.details.get("writeErrors", [])[:3])

            log.info(
                "Sheet %r: %d new, %d duplicate, %d rows skipped",
                sheet_name, inserted, matched, skipped_rows,
            )
            report.append({
                "sheet":        sheet_name,
                "inserted":     inserted,
                "duplicates":   matched,
                "skipped_rows": skipped_rows,
            })

    return report


def search_by_phone(phone: str, mongo_uri: str = MONGO_URI) -> list:
    """Return all sales records grouped by channel for a given phone number."""
    phone = normalize_phone(phone) or phone
    col = MongoClient(mongo_uri)[DB_NAME][COLLECTION]

    return list(col.aggregate([
        {"$match": {"phone": phone}},
        {"$group": {
            "_id":     "$channel",
            "count":   {"$sum": 1},
            "records": {"$push": {
                "sale_date":   "$sale_date",
                "center_name": "$center_name",
            }},
        }},
        {"$project": {
            "channel": "$_id",
            "count":    1,
            "records":  1,
            "_id":      0,
        }},
        {"$sort": {"channel": 1}},
    ]))


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_sales_excel.py <path_to_excel>")
        sys.exit(1)

    path = sys.argv[1]
    if not Path(path).exists():
        print(f"File not found: {path}")
        sys.exit(1)

    results = import_excel(path)
    print("\n── Import Summary ──")
    for r in results:
        if r.get("skipped"):
            print(f"  {r['sheet']}: SKIPPED ({r['reason']})")
        else:
            print(f"  {r['sheet']}: {r['inserted']} new | {r['duplicates']} duplicate | {r['skipped_rows']} rows without phone")
